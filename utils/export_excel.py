# utils/export_excel.py

from datetime import datetime
import os
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QMessageBox
import sqlite3  # 新增：导入 sqlite3
import traceback  # 新增：导入 traceback

class ExcelExporter:
    def __init__(self, db_manager, parent=None):
        self.db = db_manager
        self.parent = parent
        # 配置（可移到外部 config 文件）
        self.config = {
            "output_dir": "exports",
            "negative_expense_items": ["社保"],
            "columns": [
                "序号", "创建时间", "项目名称", "金额", "类型",
                "支付方式", "阶段", "状态", "备注（收入）", "支出详情（支出）"
            ]
        }

    def export_by_year(self, year, output_dir=None):
        output_dir = output_dir or self.config["output_dir"]
        start_time = datetime.now()
        
        try:
            # 创建导出目录
            if not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir)
                    logging.debug(f"创建导出目录: {output_dir}")
                except OSError as e:
                    logging.error(f"创建目录失败: {str(e)}")
                    if self.parent:
                        QMessageBox.warning(self.parent, "错误", f"无法创建目录 {output_dir}：{str(e)}")
                    return False, "dir_creation_failed"

            # 定义目标文件路径
            output_file = os.path.join(output_dir, f"{year}_transactions.xlsx")
            logging.debug(f"导出文件路径: {output_file}")

            # 初始化 ExcelWriter
            writer = None
            try:
                writer = pd.ExcelWriter(output_file, engine="openpyxl")
            except (PermissionError, IOError) as e:
                logging.error(f"初始化 ExcelWriter 失败: {str(e)}")
                if self.parent:
                    QMessageBox.warning(self.parent, "导出失败", "无法导出！请关闭文件再导出！")
                return False, "file_locked"
            except MemoryError:
                logging.error("内存不足，无法初始化 ExcelWriter")
                if self.parent:
                    QMessageBox.warning(self.parent, "错误", "内存不足，无法导出！")
                return False, "memory_error"
            except Exception as e:
                logging.error(f"初始化 ExcelWriter 未知错误: {str(e)}")
                if self.parent:
                    QMessageBox.warning(self.parent, "错误", f"初始化导出文件失败：{str(e)}")
                return False, "writer_init_failed"

            has_data = False
            negative_expense_items = self.config["negative_expense_items"]

            for month in range(1, 13):
                try:
                    transactions = self.db.get_monthly_transactions(year, month)
                    if not transactions:
                        logging.debug(f"年份 {year} 月 {month} 无交易数据，跳过")
                        continue

                    data = []
                    for idx, trans in enumerate(transactions, 1):
                        trans_id = trans[0]
                        created_at = trans[1]
                        project_name = trans[2] if trans[2] else "未知项目"
                        amount = trans[3]
                        trans_type = trans[4]
                        payment_method = trans[5]
                        stage = trans[6] if trans[6] else ""
                        status = trans[7]

                        remark = ""
                        if trans_type == "收入":
                            try:
                                with self.db.connect() as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("SELECT content FROM remarks WHERE transaction_id = ?", (trans_id,))
                                    result = cursor.fetchone()
                                    remark = result[0] if result else ""
                            except sqlite3.Error as e:
                                logging.error(f"查询 remarks 失败 (trans_id={trans_id}): {str(e)}")
                                remark = "查询失败"

                        expense_details = ""
                        if trans_type == "支出":
                            try:
                                with self.db.connect() as conn:
                                    cursor = conn.cursor()
                                    cursor.execute(
                                        "SELECT name, amount FROM expense_details WHERE transaction_id = ?",
                                        (trans_id,)
                                    )
                                    details = cursor.fetchall()
                                    if details:
                                        detail_strings = [
                                            f"{name}:{-amount:.2f}" if name in negative_expense_items else f"{name}:{amount:.2f}"
                                            for name, amount in details
                                        ]
                                        total = sum(
                                            -amount if name in negative_expense_items else amount
                                            for name, amount in details
                                        )
                                        expense_details = ", ".join(detail_strings) + f", 总额:{total:.2f}"
                            except sqlite3.Error as e:
                                logging.error(f"查询 expense_details 失败 (trans_id={trans_id}): {str(e)}")
                                expense_details = "查询失败"

                        data.append([
                            idx, created_at, project_name, f"{amount:.2f}", trans_type,
                            payment_method, stage, status, remark, expense_details
                        ])

                    if data:
                        has_data = True
                        try:
                            df = pd.DataFrame(data, columns=self.config["columns"])
                            df.to_excel(writer, sheet_name=f"{month}月", index=False)
                        except MemoryError:
                            logging.error(f"写入 {month} 月数据失败：内存不足")
                            if self.parent:
                                QMessageBox.warning(self.parent, "错误", f"写入 {month} 月数据失败：内存不足")
                            continue
                        except Exception as e:
                            logging.error(f"写入 {month} 月数据失败: {str(e)}")
                            if self.parent:
                                QMessageBox.warning(self.parent, "错误", f"写入 {month} 月数据失败：{str(e)}")
                            continue

                        worksheet = writer.sheets[f"{month}月"]
                        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        header_font = Font(bold=True)
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        for col_idx, column in enumerate(worksheet.columns, 1):
                            max_length = 0
                            column_letter = get_column_letter(col_idx)
                            for cell in column:
                                try:
                                    cell_value = str(cell.value)
                                    length = sum(2 if ord(char) > 127 else 1 for char in cell_value)
                                    max_length = max(max_length, length)
                                except:
                                    pass
                            adjusted_width = max_length * 1.2
                            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)

                        try:
                            total_income, total_expense, net_income = self.db.get_monthly_summary(year, month)
                            logging.info(f"Year: {year}, Month: {month}, Total Income: {total_income}, Total Expense: {total_expense}, Net Income: {net_income}")
                            summary_row = ["", "", "", "", "", "", "", "总收入", f"{total_income:.2f}", ""]
                            worksheet.append(summary_row)
                            worksheet.append(["", "", "", "", "", "", "", "总支出", f"{total_expense:.2f}", ""])
                            worksheet.append(["", "", "", "", "", "", "", "净收入", f"{net_income:.2f}", ""])
                            worksheet.append(["", "", "", "", "", "", "", "总额", f"{net_income:.2f}元", ""])
                        except sqlite3.Error as e:
                            logging.error(f"获取月度汇总失败 (year={year}, month={month}): {str(e)}")
                            worksheet.append(["", "", "", "", "", "", "", "汇总失败", "无法计算", ""])

                except Exception as e:
                    logging.error(f"处理月份 {month} 数据失败: {str(e)}\n{traceback.format_exc()}")
                    continue  # 跳过失败的月份，继续处理其他月份

            if has_data:
                try:
                    writer.close()
                    elapsed = (datetime.now() - start_time).total_seconds()
                    logging.info(f"成功导出 {year} 年的数据到 {output_file}，耗时 {elapsed:.2f} 秒")
                    return True, None
                except Exception as e:
                    logging.error(f"保存 Excel 文件失败: {str(e)}\n{traceback.format_exc()}")
                    if self.parent:
                        QMessageBox.warning(self.parent, "导出失败", f"无法保存文件：{str(e)}")
                    return False, "file_save_failed"
            else:
                writer.close()
                logging.info(f"{year} 年没有数据，跳过导出")
                return False, "no_data"

        except Exception as e:
            logging.error(f"导出过程未知错误: {str(e)}\n{traceback.format_exc()}")
            if writer:
                try:
                    writer.close()
                except:
                    pass
            if self.parent:
                QMessageBox.warning(self.parent, "错误", f"导出失败：{str(e)}")
            return False, "unknown_error"

    def export_all_years(self, output_dir=None):
        output_dir = output_dir or self.config["output_dir"]
        years = self.db.get_years()
        for year in years:
            self.export_by_year(year, output_dir)